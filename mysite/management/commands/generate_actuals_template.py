# mysite/management/commands/generate_actuals_template.py

"""
Produces a ready-to-use `.xlsx` file clients can download, fill in,
and upload via the API. Includes a column-header row, data-validation
dropdowns for `period_type` choices, and example data rows.

**Usage:**

```bash
# Generate template for client 'acme' with monthly periods
python manage.py generate_actuals_template --client acme --type month

# Generate template for quarterly periods, custom output path
python manage.py generate_actuals_template --client acme --type quarter --out /tmp/acme_q.xlsx

# Include 5 example rows
python manage.py generate_actuals_template --client acme --type month --examples 5
```

"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone

from mysite.models.demand.actuals import PERIOD_TYPE_CHOICES, PERIOD_FREQ_MAP


HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
EXAMPLE_FILL = PatternFill("solid", fgColor="EBF3FB")
LOCKED_FILL  = PatternFill("solid", fgColor="D9D9D9")


class Command(BaseCommand):
    help = (
        "Generate an .xlsx actuals upload template for a given client.\n"
        "Usage: python manage.py generate_actuals_template "
        "--client <client_id> [--type <period_type>] [--out <path>]"
    )

    def add_arguments(self, parser):
        parser.add_argument(
            '--client', required=True,
            help="client_id of the client to generate the template for."
        )
        parser.add_argument(
            '--type', dest='period_type', default='month',
            choices=[k for k, _ in PERIOD_TYPE_CHOICES],
            help="Period type to pre-fill (default: month)."
        )
        parser.add_argument(
            '--out', dest='output_path', default=None,
            help="Output file path. Default: actuals_template_<client>_<type>.xlsx"
        )
        parser.add_argument(
            '--examples', type=int, default=3,
            help="Number of example rows to include (default: 3)."
        )

    def handle(self, *args, **options):
        from mysite.models import Client, Item
        from mysite.models.demand.hierarchy import PlanningLocation, PlanningCustomer

        client_id   = options['client']
        period_type = options['period_type']
        output_path = options['output_path'] or (
            f"actuals_template_{client_id}_{period_type}.xlsx"
        )
        n_examples  = options['examples']

        # Resolve client
        try:
            client = Client.objects.get(client_id=client_id)
        except Client.DoesNotExist:
            raise CommandError(f"Client '{client_id}' not found.")

        # Fetch sample items, locations, customers for the example rows
        sample_items     = list(
            Item.objects.filter(client=client, status='active')
            .values_list('item_id', flat=True)[:n_examples]
        )
        sample_locations = list(
            PlanningLocation.objects.filter(client=client, is_active=True, is_leaf=True)
            .values_list('code', flat=True)[:n_examples]
        )
        sample_customers = list(
            PlanningCustomer.objects.filter(client=client, is_active=True)
            .values_list('code', flat=True)[:n_examples]
        ) + ['']   # include blank to show optional

        # ── Build workbook ────────────────────────────────────────────────────
        wb = openpyxl.Workbook()

        # ── Sheet 1: Data entry sheet ─────────────────────────────────────────
        ws = wb.active
        ws.title = "Actuals Upload"

        # Column definitions: (header, width, example_value, note)
        columns = [
            ("period_start",   18, self._example_period_start(period_type),
             "First day of the period bucket. Format: YYYY-MM-DD"),
            ("item_id",        24, sample_items[0] if sample_items else "ITEM-001",
             "Item identifier. Must be an active item for this client."),
            ("location_code",  20, sample_locations[0] if sample_locations else "LOC-001",
             "Planning location code. Must be an active leaf location."),
            ("customer_code",  20, sample_customers[0] if sample_customers else "",
             "Optional. Leave blank for unattributed demand."),
            ("qty",            14, "100",
             "Sales quantity in base UoM. Must be >= 0."),
            ("revenue",        16, "15000.00",
             "Optional. Revenue in client base currency."),
        ]

        # Write header row
        for col_idx, (col_name, width, _, _note) in enumerate(columns, start=1):
            cell               = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font          = HEADER_FONT
            cell.fill          = HEADER_FILL
            cell.alignment     = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 22

        # Write example rows
        for ex_row in range(n_examples):
            row_idx = ex_row + 2
            row_vals = [
                self._example_period_start(period_type, offset=ex_row),
                sample_items[ex_row % len(sample_items)] if sample_items else f"ITEM-00{ex_row+1}",
                sample_locations[ex_row % len(sample_locations)] if sample_locations else f"LOC-00{ex_row+1}",
                sample_customers[ex_row % len(sample_customers)],
                str((ex_row + 1) * 100),
                str((ex_row + 1) * 15000),
            ]
            for col_idx, val in enumerate(row_vals, start=1):
                cell       = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill  = EXAMPLE_FILL

        # Freeze header row
        ws.freeze_panes = "A2"

        # ── Sheet 2: Notes / Instructions ─────────────────────────────────────
        ws_notes = wb.create_sheet("Instructions")
        instructions = [
            ("Actuals Upload Template", True),
            ("", False),
            (f"Client:      {client.client_id}", False),
            (f"Period Type: {period_type}", False),
            (f"Generated:   {timezone.now():%Y-%m-%d %H:%M}", False),
            ("", False),
            ("COLUMN GUIDE", True),
            ("period_start  — YYYY-MM-DD. Must be the first day of a valid bucket.", False),
        ]
        if period_type == 'month':
            instructions.append(
                ("               Example: 2024-01-01 for January 2024.", False)
            )
        elif period_type == 'quarter':
            instructions.append(
                ("               Example: 2024-01-01 (Q1), 2024-04-01 (Q2).", False)
            )
        elif period_type == 'week':
            instructions.append(
                ("               Must be a Monday. Example: 2024-01-01.", False)
            )
        instructions += [
            ("item_id       — Must match an active item for this client.", False),
            ("location_code — Must match an active leaf planning location.", False),
            ("customer_code — Optional. Leave blank for unattributed demand.", False),
            ("qty           — Decimal number >= 0. Required.", False),
            ("revenue       — Decimal number. Optional.", False),
            ("", False),
            ("RULES", True),
            ("1. Do not change column headers.", False),
            ("2. Delete example rows before uploading.", False),
            ("3. Uploading the same file twice updates (not duplicates) existing rows.", False),
            ("4. Rows with errors are skipped; other rows are still imported.", False),
            ("5. Check the import status API for row-level error details.", False),
        ]

        for row_idx, (text, is_heading) in enumerate(instructions, start=1):
            cell = ws_notes.cell(row=row_idx, column=1, value=text)
            if is_heading:
                cell.font = Font(bold=True, size=12)
        ws_notes.column_dimensions['A'].width = 75

        # ── Sheet 3: Valid reference values ───────────────────────────────────
        ws_ref = wb.create_sheet("Reference Values")

        ref_cols = [
            ("Valid Item IDs",
             list(Item.objects.filter(client=client, status='active')
                  .values_list('item_id', flat=True)[:200])),
            ("Valid Location Codes",
             list(PlanningLocation.objects.filter(client=client, is_active=True, is_leaf=True)
                  .values_list('code', flat=True)[:200])),
            ("Valid Customer Codes (optional)",
             list(PlanningCustomer.objects.filter(client=client, is_active=True)
                  .values_list('code', flat=True)[:200])),
        ]

        for col_idx, (heading, values) in enumerate(ref_cols, start=1):
            cell       = ws_ref.cell(row=1, column=col_idx, value=heading)
            cell.font  = Font(bold=True)
            cell.fill  = HEADER_FILL
            cell.font  = HEADER_FONT
            ws_ref.column_dimensions[get_column_letter(col_idx)].width = 30
            for row_idx, val in enumerate(values, start=2):
                ws_ref.cell(row=row_idx, column=col_idx, value=val)

        # ── Save ──────────────────────────────────────────────────────────────
        wb.save(output_path)
        self.stdout.write(
            self.style.SUCCESS(f"Template written to: {output_path}")
        )

    def _example_period_start(self, period_type: str, offset: int = 0) -> str:
        """Return an example period_start string for the given period_type."""
        import datetime
        from dateutil.relativedelta import relativedelta
        base = datetime.date(2025, 1, 1)  # always use a clean anchor
        if period_type == 'week':
            d = base + datetime.timedelta(weeks=offset)
        elif period_type == 'month':
            d = base + relativedelta(months=offset)
        elif period_type == 'bimonth':
            d = base + relativedelta(months=offset * 2)
        elif period_type == 'quarter':
            d = base + relativedelta(months=offset * 3)
        elif period_type == 'halfyear':
            d = base + relativedelta(months=offset * 6)
        elif period_type == 'year':
            d = base + relativedelta(years=offset)
        else:
            d = base + datetime.timedelta(days=offset)
        return d.isoformat()
    
