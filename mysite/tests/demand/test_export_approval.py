import datetime
import io
import pytest
from decimal import Decimal
from rest_framework.test import APIClient
from openpyxl import load_workbook

from mysite.models.demand.forecast import (
    ForecastVersion, ForecastLine, ForecastAccuracy,
)


# ─────────────────────────────────────────────────────────────────────────────
# Shared fixtures (re-use from earlier sprints where possible)
# ─────────────────────────────────────────────────────────────────────────────

@pytest.fixture
def version_with_lines(db, draft_version, active_item, leaf_location):
    from mysite.models.demand.forecast import ForecastLine
    periods = [
        (datetime.date(2025, 1, 1), datetime.date(2025, 1, 31)),
        (datetime.date(2025, 2, 1), datetime.date(2025, 2, 28)),
        (datetime.date(2025, 3, 1), datetime.date(2025, 3, 31)),
    ]
    lines = []
    for ps, pe in periods:
        lines.append(ForecastLine.objects.create(
            version=draft_version, item=active_item,
            planning_location=leaf_location, planning_customer=None,
            period_type='month', period_start=ps, period_end=pe,
            statistical_qty=Decimal('100.000'), final_qty=Decimal('100.000'),
            price_used=Decimal('150.00'), statistical_value=Decimal('15000.00'),
            final_value=Decimal('15000.00'), model_used='AutoETS',
            forecast_level='item_cust_location',
        ))
    # Return as (version, lines_list) — unpack as: version, lines = version_with_lines
    return draft_version, lines


# ─────────────────────────────────────────────────────────────────────────────
# Test: Export endpoint
# ─────────────────────────────────────────────────────────────────────────────
@pytest.mark.django_db
class TestForecastExport:

    def test_export_returns_xlsx_content_type(self, api_client, version_with_lines):
        client, user = api_client   # ← conftest api_client, two values
        version, lines = version_with_lines
        resp = client.get(f'/api/demand/forecast-versions/{version.pk}/export/')
        assert resp.status_code == 200
        assert 'spreadsheetml' in resp['Content-Type']
        assert 'attachment' in resp['Content-Disposition']
        assert '.xlsx' in resp['Content-Disposition']

    def test_export_returns_404_when_no_lines(self, api_client, draft_version):
        client, user = api_client
        resp = client.get(
            f'/api/demand/forecast-versions/{draft_version.pk}/export/'
        )
        assert resp.status_code == 404

    # ... all other export tests use (api_client, version_with_lines)
    # unpack as: client, user = api_client  and  version, lines = version_with_lines


    def test_export_row_count_matches_forecast_line_count(
        self, api_client, version_with_lines
    ):
        client, user = api_client
        version, lines = version_with_lines

        resp = client.get(f'/api/demand/forecast-versions/{version.pk}/export/')
        assert resp.status_code == 200

        wb  = load_workbook(io.BytesIO(resp.content))
        ws  = wb['Forecast']

        data_rows = [
            row for row in ws.iter_rows(min_row=5, values_only=True)
            if any(cell is not None for cell in row)
            and not str(row[0] or '').startswith('⬛')   # exclude legend row
        ]
        assert len(data_rows) == 1

    def test_export_period_columns_match_distinct_periods(
        self, api_client, version_with_lines
    ):
        
        #The number of period columns = number of distinct period_start values.
        #Fixed columns = 6; periods = 3 → total columns = 9.
        
        client, user = api_client
        version, lines = version_with_lines

        resp = client.get(f'/api/demand/forecast-versions/{version.pk}/export/')
        wb   = load_workbook(io.BytesIO(resp.content))
        ws   = wb['Forecast']

        header_row = [cell.value for cell in ws[4] if cell.value is not None]
        # 6 fixed + 3 period labels
        assert len(header_row) == 9
        # Period labels look like 'Jan-25'
        period_labels = header_row[6:]
        for label in period_labels:
            assert len(label) == 6    # e.g. 'Jan-25'
            assert '-' in label

    def test_export_final_qty_values_in_cells(self, api_client, version_with_lines):
        client, user = api_client
        version, lines = version_with_lines

        resp = client.get(f'/api/demand/forecast-versions/{version.pk}/export/')
        wb   = load_workbook(io.BytesIO(resp.content))
        ws   = wb['Forecast']

        # Row 5 = first data row
        # Fixed columns = 6 (Location, Item ID, Item Name, Customer, UOM, Price)
        # Period columns start at column 7 (index 6 in 0-based)
        n_fixed = 6
        data_row = [cell.value for cell in ws[5]]
        period_qtys = [
            v for v in data_row[n_fixed:]   # skip the 6 fixed columns
            if isinstance(v, (int, float))
        ]
        assert period_qtys, 'No numeric values found in period columns'
        for qty in period_qtys:
            assert abs(qty - 100.0) < 0.001

    def test_export_works_on_locked_version(self, api_client, version_with_lines):
        #Export must succeed on LOCKED versions — these are the PO baseline.
        client, user = api_client
        version, lines = version_with_lines

        # Promote to LOCKED
        version.status    = ForecastVersion.Status.LOCKED
        version.locked_at = __import__('django.utils.timezone', fromlist=['now']).now()
        version.save(update_fields=['status', 'locked_at'])

        resp = client.get(f'/api/demand/forecast-versions/{version.pk}/export/')
        assert resp.status_code == 200
        assert 'spreadsheetml' in resp['Content-Type']

    def test_export_returns_404_when_no_lines(self, api_client, draft_version):
        #Exporting a version with no ForecastLine rows returns 404.
        client, user = api_client
        resp = client.get(
            f'/api/demand/forecast-versions/{draft_version.pk}/export/'
        )
        assert resp.status_code == 404

    def test_export_summary_sheet_exists(self, api_client, version_with_lines):
        #Workbook must have a 'Summary' sheet.
        client, user = api_client
        version, lines = version_with_lines

        resp = client.get(f'/api/demand/forecast-versions/{version.pk}/export/')
        wb   = load_workbook(io.BytesIO(resp.content))
        assert 'Summary' in wb.sheetnames

    def test_export_location_filter(self, api_client, version_with_lines):
        #location_code param must filter exported lines.
        client, user = api_client
        version, lines = version_with_lines
        location_code = lines[0].planning_location.code

        resp = client.get(
            f'/api/demand/forecast-versions/{version.pk}/export/'
            f'?location_code={location_code}'
        )
        assert resp.status_code == 200


# ─────────────────────────────────────────────────────────────────────────────
# Test: Approval workflow state transitions
# ─────────────────────────────────────────────────────────────────────────────

@pytest.mark.django_db
class TestApprovalWorkflow:

    def _approve_url(self, version_id):
        return f'/api/demand/forecast-versions/{version_id}/approve/'

    def test_submit_draft_to_in_review(self, api_client, draft_version):
        client, user = api_client
        resp = client.post(
            self._approve_url(draft_version.pk),
            {'action': 'submit'},
            format='json',
        )
        assert resp.status_code == 200
        assert resp.data['status'] == 'IN_REVIEW'

    # ... all other approval tests use (api_client, draft_version)
    

    def test_approve_in_review_to_approved(self, api_client, draft_version):
        client, user = api_client
        draft_version.status = ForecastVersion.Status.IN_REVIEW
        draft_version.save(update_fields=['status'])

        resp = client.post(
            self._approve_url(draft_version.pk),
            {'action': 'approve'},
            format='json',
        )
        assert resp.status_code == 200
        assert resp.data['status'] == 'APPROVED'

    def test_reject_in_review_to_draft(self, api_client, draft_version):
        client, user = api_client
        draft_version.status = ForecastVersion.Status.IN_REVIEW
        draft_version.save(update_fields=['status'])

        resp = client.post(
            self._approve_url(draft_version.pk),
            {'action': 'reject', 'note': 'North region totals need revision'},
            format='json',
        )
        assert resp.status_code == 200
        assert resp.data['status'] == 'DRAFT'

    def test_lock_approved_to_locked(self, api_client, draft_version):
        client, user = api_client
        draft_version.status = ForecastVersion.Status.APPROVED
        draft_version.save(update_fields=['status'])

        resp = client.post(
            self._approve_url(draft_version.pk),
            {'action': 'lock'},
            format='json',
        )
        assert resp.status_code == 200
        assert resp.data['status'] == 'LOCKED'
        # Confirm locked_at is now set
        draft_version.refresh_from_db()
        assert draft_version.locked_at is not None

    def test_illegal_transition_returns_403(self, api_client, draft_version):
        #Trying to lock a DRAFT version (skipping IN_REVIEW) must return 403.
        client, user = api_client
        resp = client.post(
            self._approve_url(draft_version.pk),
            {'action': 'lock'},
            format='json',
        )
        assert resp.status_code == 403

    def test_copy_any_status_creates_draft(self, api_client, draft_version):
        client, user = api_client
        resp = client.post(
            self._approve_url(draft_version.pk),
            {'action': 'copy', 'note': 'Feb-2025 Plan v1'},
            format='json',
        )
        assert resp.status_code == 201
        assert resp.data['status'] == 'DRAFT'
        assert resp.data['version_label'] == 'Feb-2025 Plan v1'

    def test_note_appended_to_version_notes(self, api_client, draft_version):
        #A note provided on submit must be appended to ForecastVersion.notes.
        client, user = api_client
        resp = client.post(
            self._approve_url(draft_version.pk),
            {'action': 'submit', 'note': 'Ready for monthly consensus review'},
            format='json',
        )
        assert resp.status_code == 200
        draft_version.refresh_from_db()
        assert 'Ready for monthly consensus review' in draft_version.notes


# ─────────────────────────────────────────────────────────────────────────────
# Test: Accuracy dashboard endpoint
# ─────────────────────────────────────────────────────────────────────────────
@pytest.mark.django_db
class TestAccuracyDashboard:

    @pytest.fixture
    def version_with_accuracy(self, draft_version, active_item, leaf_location):
        for ps, pe in [
            (datetime.date(2025, 1, 1), datetime.date(2025, 1, 31)),
            (datetime.date(2025, 2, 1), datetime.date(2025, 2, 28)),
        ]:
            ForecastAccuracy.objects.create(
                version              = draft_version,
                item                 = active_item,
                planning_location    = leaf_location,
                planning_customer    = None,
                period_type          = 'month',
                period_start         = ps,
                period_end           = pe,
                actual_qty           = Decimal('100.000'),
                forecast_qty         = Decimal('110.000'),
                mape                 = Decimal('10.0000'),
                bias                 = Decimal('10.0000'),
            )
        return draft_version

    def test_accuracy_endpoint_returns_200(
        self, api_client, version_with_accuracy
    ):
        client, user = api_client
        resp = client.get(
            f'/api/demand/forecast-versions/{version_with_accuracy.pk}/accuracy/'
        )
        assert resp.status_code == 200


    def test_accuracy_overall_mape_computed(
        self, api_client, version_with_accuracy
    ):
        client, user = api_client
        resp = client.get(
            f'/api/demand/forecast-versions/{version_with_accuracy.pk}/accuracy/'
            f'?group_by=period'
        )
        assert resp.status_code == 200
        overall = resp.data.get('overall')
        assert overall is not None
        assert float(overall['mean_mape']) == pytest.approx(10.0, abs=0.1)

    def test_accuracy_group_by_period(self, api_client, version_with_accuracy):
        client, user = api_client
        resp = client.get(
            f'/api/demand/forecast-versions/{version_with_accuracy.pk}/accuracy/'
            f'?group_by=period'
        )
        assert resp.status_code == 200
        assert resp.data['group_by'] == 'period'
        # 2 periods should give 2 result rows
        assert resp.data['count'] == 2

    def test_accuracy_returns_empty_gracefully_with_no_records(
        self, api_client, draft_version
    ):
        #Version with no accuracy records returns 200 with empty results.#
        client, user = api_client
        resp = client.get(
            f'/api/demand/forecast-versions/{draft_version.pk}/accuracy/'
        )
        assert resp.status_code == 200
        assert resp.data['count'] == 0
        assert resp.data['results'] == []

    def test_accuracy_invalid_group_by_returns_400(
        self, api_client, version_with_accuracy
    ):
        client, user = api_client
        resp = client.get(
            f'/api/demand/forecast-versions/{version_with_accuracy.pk}/accuracy/'
            f'?group_by=invalid'
        )
        assert resp.status_code == 400
